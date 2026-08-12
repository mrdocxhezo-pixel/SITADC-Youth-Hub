"""Excel (XLSX) renderer for the Export Engine.

Built on ``openpyxl``.  Applies formula-injection hardening to every cell,
formats the header row, adds an auto-filter and a bold metadata block, and
sanitizes the worksheet name.
"""

from __future__ import annotations

import io
from typing import Any

from ..constants import ExportFormat
from ..utils import neutralize_spreadsheet_value, sanitize_sheet_name
from .base import BaseRenderer, ExportDataset, RendererRegistry, RenderResult


@RendererRegistry.register(ExportFormat.XLSX)
class XLSXRenderer(BaseRenderer):
    """Render an ExportDataset to a styled Excel workbook."""

    format = ExportFormat.XLSX

    def render(
        self, dataset: ExportDataset, configuration: Any
    ) -> RenderResult:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sanitize_sheet_name(dataset.title, default="Export")

        navy = "003366"
        light_grey = "EEF2F7"
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
        meta_font = Font(bold=True, size=10, color="444444")

        row_index = 1

        # Metadata block
        meta_rows = list(dataset.meta_rows)
        if dataset.reference:
            meta_rows.append(("Reference", dataset.reference))
        if dataset.report_period:
            meta_rows.append(("Reporting Period", dataset.report_period))
        if dataset.status:
            meta_rows.append(("Status", dataset.status))
        meta_rows.append(("Confidentiality", str(dataset.confidentiality)))
        for label, value in meta_rows:
            sheet.cell(row=row_index, column=1, value=label).font = meta_font
            sheet.cell(row=row_index, column=2, value=value)
            row_index += 1

        # Gap before header
        if meta_rows:
            row_index += 1

        # Header row
        headers = dataset.column_labels
        header_row = row_index
        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=col_index, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        row_index += 1

        # Body rows
        first_data_row = row_index
        for row in dataset.rows:
            for col_index, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_index, column=col_index)
                cell.value = neutralize_spreadsheet_value(value)
            row_index += 1

        last_data_row = max(first_data_row, row_index - 1)

        # Banded rows
        banded_fill = PatternFill(
            start_color=light_grey, end_color=light_grey, fill_type="solid"
        )
        for data_row in range(first_data_row, last_data_row + 1):
            if (data_row - first_data_row) % 2 == 1:
                for col_index in range(1, len(headers) + 1):
                    sheet.cell(row=data_row, column=col_index).fill = banded_fill

        # Auto-filter + freeze panes
        if headers:
            last_col = get_column_letter(len(headers))
            if last_data_row >= header_row:
                sheet.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"
                sheet.freeze_panes = f"A{header_row + 1}"

        # Column widths (bounded)
        for col_index in range(1, len(headers) + 1):
            width = min(
                32,
                max(
                    12,
                    len(str(headers[col_index - 1])) + 6,
                ),
            )
            sheet.column_dimensions[get_column_letter(col_index)].width = width

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return RenderResult(
            content=buffer.getvalue(),
            mime_type=self.content_type,
            filename=self.default_extension,
        )
