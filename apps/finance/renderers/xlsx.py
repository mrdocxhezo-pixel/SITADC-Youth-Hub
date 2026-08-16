"""Finance Engine Excel renderer."""

from __future__ import annotations

import io
from typing import Any, Dict, List, Union

from django.http import HttpResponse
from django.utils import timezone

# Try to import openpyxl or xlsxwriter for Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    try:
        import xlsxwriter
        XLSXWRITER_AVAILABLE = True
    except ImportError:
        XLSXWRITER_AVAILABLE = False

from .base import BaseFinanceRenderer


class FinanceExcelRenderer(BaseFinanceRenderer):
    """Finance report Excel renderer."""

    def render(self) -> bytes:
        """
        Render the data as Excel.

        Returns:
            bytes: The rendered Excel data.
        """
        if OPENPYXL_AVAILABLE:
            return self._render_with_openpyxl()
        elif XLSXWRITER_AVAILABLE:
            return self._render_with_xlsxwriter()
        else:
            # Fallback to CSV representation
            return self._render_fallback_csv()

    def _render_with_openpyxl(self) -> bytes:
        """
        Render Excel using openpyxl.

        Returns:
            bytes: The rendered Excel data.
        """
        # Create workbook and select active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Report"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        title_font = Font(bold=True, size=16)
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = self.data.get('title', 'Financial Report')
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        
        # Add timestamp
        ws.merge_cells('A2:D2')
        timestamp_cell = ws['A2']
        timestamp_cell.value = f"Generated on: {self.timestamp.strftime('%Y-%m-%d %H:%M:%S')}"
        timestamp_cell.alignment = Alignment(horizontal="center")
        
        # Add data starting from row 4
        row = 4
        
        # Add section headers and data
        for section_name, section_data in self.data.items():
            if section_name in ['title']:  # Skip metadata
                continue
                
            # Section header
            ws.merge_cells(f'A{row}:D{row}')
            header_cell = ws[f'A{row}']
            header_cell.value = section_name.replace('_', ' ').title()
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = header_alignment
            row += 1
            
            # Add section data
            if isinstance(section_data, dict):
                # Add column headers
                ws.cell(row=row, column=1, value="Key").font = header_font
                ws.cell(row=row, column=1).fill = header_fill
                ws.cell(row=row, column=2, value="Value").font = header_font
                ws.cell(row=row, column=2).fill = header_fill
                row += 1
                
                # Add data rows
                for key, value in section_data.items():
                    ws.cell(row=row, column=1, value=str(key))
                    ws.cell(row=row, column=2, value=str(value))
                    row += 1
            elif isinstance(section_data, list):
                if section_data and isinstance(section_data[0], dict):
                    # List of dictionaries - treat as table
                    if section_data:
                        # Add headers
                        headers = list(section_data[0].keys())
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=row, column=col, value=header.replace('_', ' ').title())
                            cell.font = header_font
                            cell.fill = header_fill
                        row += 1
                        
                        # Add data rows
                        for item in section_data:
                            for col, key in enumerate(headers, 1):
                                ws.cell(row=row, column=col, value=str(item.get(key, '')))
                            row += 1
                else:
                    # Simple list
                    for item in section_data:
                        ws.cell(row=row, column=1, value=str(item))
                        row += 1
            else:
                # Simple value
                ws.cell(row=row, column=1, value=str(section_data))
                row += 1
            
            # Add blank row between sections
            row += 1
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file.getvalue()

    def _render_with_xlsxwriter(self) -> bytes:
        """
        Render Excel using XlsxWriter.

        Returns:
            bytes: The rendered Excel data.
        """
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("Financial Report")
        
        # Define formats
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#366092',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        cell_format = workbook.add_format({
            'border': 1
        })
        
        # Add title
        worksheet.merge_range('A1:D1', self.data.get('title', 'Financial Report'), title_format)
        
        # Add timestamp
        worksheet.merge_range('A2:D2', f"Generated on: {self.timestamp.strftime('%Y-%m-%d %H:%M:%S')}", workbook.add_format({'align': 'center'}))
        
        # Add data starting from row 4
        row = 3  # Zero-indexed
        
        # Add section headers and data
        for section_name, section_data in self.data.items():
            if section_name in ['title']:  # Skip metadata
                continue
                
            # Section header
            worksheet.merge_range(f'A{row+1}:D{row+1}', section_name.replace('_', ' ').title(), header_format)
            row += 1
            
            # Add section data
            if isinstance(section_data, dict):
                # Add column headers
                worksheet.write(row, 0, "Key", header_format)
                worksheet.write(row, 1, "Value", header_format)
                row += 1
                
                # Add data rows
                for key, value in section_data.items():
                    worksheet.write(row, 0, str(key), cell_format)
                    worksheet.write(row, 1, str(value), cell_format)
                    row += 1
            elif isinstance(section_data, list):
                if section_data and isinstance(section_data[0], dict):
                    # List of dictionaries - treat as table
                    if section_data:
                        # Add headers
                        headers = list(section_data[0].keys())
                        for col, header in enumerate(headers):
                            worksheet.write(row, col, header.replace('_', ' ').title(), header_format)
                        row += 1
                        
                        # Add data rows
                        for item in section_data:
                            for col, key in enumerate(headers):
                                worksheet.write(row, col, str(item.get(key, '')), cell_format)
                            row += 1
                else:
                    # Simple list
                    for item in section_data:
                        worksheet.write(row, 0, str(item), cell_format)
                        row += 1
            else:
                # Simple value
                worksheet.write(row, 0, str(section_data), cell_format)
                row += 1
            
            # Add blank row between sections
            row += 1
        
        workbook.close()
        excel_data = output.getvalue()
        output.close()
        
        return excel_data

    def _render_fallback_csv(self) -> bytes:
        """
        Render fallback CSV representation when Excel libraries are not available.

        Returns:
            bytes: The rendered CSV data.
        """
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Add title
        writer.writerow([self.data.get('title', 'Financial Report')])
        writer.writerow([f"Generated on: {self.timestamp.strftime('%Y-%m-%d %H:%M:%S')}"])
        writer.writerow([])  # Empty row
        
        # Add data
        for section_name, section_data in self.data.items():
            if section_name in ['title']:  # Skip metadata
                continue
                
            writer.writerow([section_name.replace('_', ' ').title()])
            
            if isinstance(section_data, dict):
                writer.writerow(["Key", "Value"])
                for key, value in section_data.items():
                    writer.writerow([str(key), str(value)])
            elif isinstance(section_data, list):
                if section_data and isinstance(section_data[0], dict):
                    # List of dictionaries
                    headers = list(section_data[0].keys())
                    writer.writerow([h.replace('_', ' ').title() for h in headers])
                    for item in section_data:
                        writer.writerow([str(item.get(h, '')) for h in headers])
                else:
                    # Simple list
                    for item in section_data:
                        writer.writerow([str(item)])
            else:
                # Simple value
                writer.writerow(["Value", str(section_data)])
            
            writer.writerow([])  # Empty row between sections
        
        csv_content = output.getvalue()
        output.close()
        
        return csv_content.encode('utf-8')